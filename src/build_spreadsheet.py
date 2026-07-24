"""Turn extracted JSON records into a formatted .xlsx workbook - one tab
per source PDF, low-confidence rows highlighted with a review comment."""
import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill

from schema import XLSX_HEADERS

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF")
FLAG_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
NEPALI_FONT = Font(name="Nirmala UI")  # renders Devanagari correctly in Excel
NORMAL_FONT = Font(name="Arial")


def _write_sheet(ws, records: list[dict]) -> int:
    ws.append(XLSX_HEADERS)
    for col in range(1, len(XLSX_HEADERS) + 1):
        c = ws.cell(row=1, column=col)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")

    flagged = 0
    for r, rec in enumerate(records, start=2):
        values = [
            rec.get("wp", ""), rec.get("date", ""),
            rec.get("latitude", ""), rec.get("longitude", ""), rec.get("elevation_m", ""),
            rec.get("species", ""), rec.get("count", ""),
            rec.get("notes_nepali", ""), rec.get("notes_english", ""),
        ]
        for col, val in enumerate(values, start=1):
            ws.cell(row=r, column=col, value=val).font = NORMAL_FONT
        ws.cell(row=r, column=8).font = NEPALI_FONT

        flag = rec.get("flag")
        if flag:
            flagged += 1
            for col in range(1, len(XLSX_HEADERS) + 1):
                ws.cell(row=r, column=col).fill = FLAG_FILL
            ws.cell(row=r, column=1).comment = Comment(str(flag), "Transcription review")

    widths = [14, 12, 13, 13, 12, 14, 8, 34, 46]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    return flagged


def build_workbook(records_by_source: dict[str, list[dict]], output_path: str):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    summary_lines = ["Source file | Rows | Flagged for review", "-" * 50]
    for source, records in records_by_source.items():
        # Excel sheet names: <=31 chars, no []:*?/\
        safe_name = "".join(c for c in source if c not in r'[]:*?/\\')[:31]
        ws = wb.create_sheet(safe_name)
        flagged = _write_sheet(ws, records)
        summary_lines.append(f"{source} | {len(records)} | {flagged}")

    readme = wb.create_sheet("Read Me", 0)
    for i, line in enumerate([
        "Each tab = one source PDF. Columns match schema.py / prompts/extraction_prompt.md.",
        "",
        "Species / Count are split out from notebook shorthand (e.g. 'Blue sheep-19' -> ",
        "Species: Blue sheep, Count: 19). Blank Species = logistics-only entry, no wildlife seen.",
        "",
        "Notes (Nepali - original): verbatim Devanagari script, no translation folded in.",
        "Notes (English): description + a labelled DRAFT translation of any Nepali note -",
        "correct these, don't trust them as final.",
        "",
        "Yellow rows have a comment on the W.P. cell explaining what needs a human check.",
        "",
    ] + summary_lines, start=1):
        readme.cell(row=i, column=1, value=line).font = NORMAL_FONT
    readme.column_dimensions["A"].width = 100

    wb.save(output_path)
